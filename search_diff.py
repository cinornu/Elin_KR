import os
import pathlib
from openpyxl import load_workbook
from distutils.dir_util import copy_tree

print("\n* 알림: Dialog쪽 파일에서는 새로 추가/삭제된 행을 찾는 것이 어려울 수 있습니다. 공백 검색기 등과 병행해 주세요.\n")

# 폴더 경로 설정
# Elin_KR 폴더에 있는 것을 가정
path_dir = str(pathlib.Path(__file__).parent.resolve())

path_detail = ["\\Mod_Korean\\Lang\\KR\\Data", "\\Mod_Korean\\Lang\\KR\\Dialog", "\\Mod_Korean\\Lang\\KR\\Dialog\\Drama", "\\Mod_Korean\\Lang\\KR\\Game"]
prev_folder = path_dir + "\\DIFF_PREVIOUS"

# 비교군 (DIFF_PREVIOUS 폴더) 가 없다면 새로 생성 후 파일 복제
if not os.path.isdir('DIFF_PREVIOUS'):
    yesno = input("경고: 비교 대상인 과거판 파일이 존재하지 않습니다. 현재 파일을 복사하여 새로 생성합니까? (y/n): ")
    if yesno == "y" or yesno == "Y":
        os.mkdir("DIFF_PREVIOUS")
        os.mkdir("DIFF_PREVIOUS\\Mod_Korean")

        copy_tree(path_dir + "\\Mod_Korean", prev_folder + "\\Mod_Korean")

        print("\n비교 대상 파일을 새로 생성했습니다 (DIFF_PREVIOUS 디렉토리). 프로그램을 종료합니다.")
        exit()
    else:
        exit()

# 버전 파일 위치
path_file_version = path_dir + "\\Mod_Korean\\Lang\\KR\\version.ini"
path_file_version_prev = prev_folder + "\\Mod_Korean\\Lang\\KR\\version.ini"

# 파일 버전 가져오기
file_version = open(path_file_version, "r", encoding="utf-8")
version = file_version.read()
version = version[-6:-1]
file_version.close()

file_version = open(path_file_version_prev, "r", encoding="utf-8")
version_prev = file_version.read()
version_prev = version_prev[-6:-1]
file_version.close()

# 버전 체크; 서로 같으면 Y/N
if version == version_prev:
    yesno = input(f"경고: 과거판과 현재판의 버전이 서로 같습니다 ({version}). 그래도 검색합니까? (y/n): ")
    if yesno == "y" or yesno == "Y":
        pass
    else:
        exit()
else:
    yesno = input(f"과거판 버전: {version_prev}\n현재판 버전: {version}\n검색을 시작합니까? (y/n): ")
    if yesno == "y" or yesno == "Y":
        pass
    else:
        exit()

# 각 폴더 경로 지정
array_path = []
array_path_prev = []

for i in range(0, len(path_detail)):
    array_path.append(path_dir + path_detail[i])
    array_path_prev.append(prev_folder + path_detail[i])

# 각 파일들의 경로를 저장할 배열
array_file = []
array_file_prev = []

# 각 파일들의 이름도 따로 저장 (비교용)
array_filename = []
array_filename_prev = []

# 결과 저장 리스트
results = []

# 하위 폴더들 내 모든 엑셀 파일 순회
# 과거판
for folder_path in array_path_prev:
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):  # ~$ 임시 파일 제외
            file_path = os.path.join(folder_path, filename)
            print(f"과거판 파일 취득 중: {filename}")  # 현재 파일명 출력
            array_file_prev.append(file_path)
            array_filename_prev.append(filename)

# 최신판
for folder_path in array_path:
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            file_path = os.path.join(folder_path, filename)
            print(f"최신판 파일 취득 중: {filename}")
            array_file.append(file_path)
            array_filename.append(filename)

print("파일 체크 중...")

# 두 파일 배열의 원소 중 다른 것이 있음 = 새로 추가/제거된 파일이 존재함
filenum_curr = len(array_filename)
filenum_prev = len(array_filename_prev)

if filenum_curr != filenum_prev:
    target_file = []
    target_status = ""
    
    # 최신판 파일이 더 많음; 과거판 파일명들을 배열에서 제거하여 새 파일 이름 특정
    if filenum_curr > filenum_prev:
        for name in array_filename_prev:
            array_filename.remove(name)
        for t in array_filename:
            target_file.append(t)
        target_status = "(최신판에 새로 추가됨)"
    # 과거판 파일이 더 많음
    else:
        for name in array_filename:
            array_filename_prev.remove(name)
        for t in array_filename_prev:
            target_file.append(t)
        target_status = "(최신판에서 삭제됨)"

    print(f"\n!! 경고 !!\n과거 파일과 현재 파일의 개수가 다릅니다. 최신판에 새로 추가/제거된 파일이 있습니다.\n해당 파일: {target_file} {target_status}.\n")
    print("Diff 검색은 과거판/최신판의 파일 개수가 동일해야만 수행할 수 있습니다.")
    print("새로 추가/제거된 파일에 대한 작업을 끝내고, 수동으로 과거판 파일을 최신화해 주십시오.")
    exit()


# 모든 파일에 대해서 비교 실시; 파일 개수가 같은 게 확인됐으므로, 같은 index의 파일명은 항상 동일함
for i in range(0, len(array_file)):
    workbook_prev = load_workbook(array_file_prev[i], data_only=True)
    workbook_curr = load_workbook(array_file[i], data_only=True)

    # 시트 개수가 같은지 체크; 같지 않으면 중단
    if len(workbook_curr.sheetnames) != len(workbook_prev.sheetnames):
        print(f"!! 경고 !!\n과거판 파일과 현재판 파일의 시트 개수가 다릅니다. 특정 파일에 새로 추가/제거된 시트가 있습니다.\n해당 파일: {array_filename[i]}.")
        print("Diff 검색은 과거판/최신판 파일의 시트 개수가 동일해야만 수행할 수 있습니다.")
        print("수정된 파일에 대한 작업을 끝내고, 수동으로 과거판 파일을 최신화해 주십시오.")
        exit()

    print(f"파일 비교 중: {array_filename[i]}")

    # 각 시트별로 검색
    for sheet_name in workbook_curr.sheetnames:
        sheet_curr = workbook_curr[sheet_name]
        sheet_prev = workbook_prev[sheet_name]

        # 각 행에 대해 텍스트 비교
        row_curr = list(sheet_curr.iter_rows(values_only=True))
        row_curr_for_check = list(sheet_curr.iter_rows(values_only=True))
        row_prev = list(sheet_prev.iter_rows(values_only=True))

        # 현재 파일 행 - 과거 파일 행 을 시행해서, 남는 것이 차이점임
        row_for_diff = []
        for row in row_prev:
            if row in row_curr:
                row_curr.remove(row)
            else:
                row_for_diff.append(row)

        # 새 행이 추가되거나 기존 행이 삭제되었을 경우 처리
        row_new = []
        row_deleted = []

        while(len(row_curr) != len(row_for_diff)):
            # 새 행이 추가되었을 때 -> prev에 해당 행의 id가 있는지 체크
            # 만약 dialog 관련 파일일 경우, 9행과 10행 (id, text_jp) 기준 체크
            # dialog의 경우, id가 없는 행 (대사가 없고 스크립트 명령만 있는 행) 은 자동 제외
            if len(row_curr) > len(row_for_diff):
                for target in row_curr:
                    founded = False
                    for p in row_prev:
                        if "Drama" not in array_file[i] and target[0] == p[0]:
                            founded = True
                            break
                        elif "Drama" in array_file[i] and target[8] == p[8] and target[9] == p[9] and target[9] != None:
                            founded = True
                            break
                    # id를 prev에서 못찾았음 = 이 행이 새로 추가된 행임
                    if not founded:
                        # dialog 파일의 스크립트 행은 제외 (새 행 배열에 추가하지 않음)
                        if "Drama" in array_file[i] and target[8] == None:
                            row_curr.remove(target)
                        else:
                            # 새 행 배열에 추가하고 기존 배열에서 제거
                            row_new.append(target)
                            row_curr.remove(target)
            # 기존 행이 제거되었을 때 -> curr에 해당 행의 id가 있는지 체크
            if len(row_curr) < len(row_for_diff):
                for target in row_for_diff:
                    founded = False
                    for p in row_curr_for_check:
                        if "Drama" not in array_file[i] and target[0] == p[0]:
                            founded = True
                            break
                        elif "Drama" in array_file[i] and target[8] == p[8] and target[9] == p[9] and target[9] != None:
                            founded = True
                            break
                    # id를 curr에서 못찾음 = 이 행은 삭제된 행임
                    if not founded:
                        # dialog 파일의 스크립트 행은 제외 (제거된 행 배열에 추가하지 않음)
                        if "Drama" in array_file[i] and target[8] == None:
                            row_for_diff.remove(target)
                        else:
                            # 제거된 행 배열에 추가하고 기존 배열에서 제거
                            row_deleted.append(target)
                            row_for_diff.remove(target)

        # 현재판 파일에서 행수 탐색하여 저장
        arr_result_change = []
        arr_result_change_diff = []
        arr_result_new = []
        arr_result_deleted = []

        if row_curr != []:
            for row in row_curr:
                num = row_curr_for_check.index(row) + 1
                arr_result_change.append((num, row))
            # 결과물 행수 기준 정렬
            arr_result_change.sort(key = lambda x:x[0])

        if row_for_diff != []:
            for row in row_for_diff:
                num = row_prev.index(row) + 1
                arr_result_change_diff.append((num, row))
            arr_result_change_diff.sort(key = lambda x:x[0])

        if row_new != []:
            for row in row_new:
                num = row_curr_for_check.index(row) + 1
                arr_result_new.append((num, row))
            arr_result_new.sort(key = lambda x:x[0])

        if row_deleted != []:
            for row in row_deleted:
                num = row_prev.index(row) + 1
                arr_result_deleted.append((num, row))
            arr_result_deleted.sort(key = lambda x:x[0])

        # 결과 저장
        results.append((array_file[i], sheet_name, arr_result_change, arr_result_change_diff, arr_result_new, arr_result_deleted, array_filename[i]))

# 상세 결과 텍스트 파일들을 저장할 경로 지정
result_folder = path_dir + "\\diff_detail" + f"_{version}"
if not os.path.isdir(f'diff_detail_{version}'):
    os.mkdir(f"diff_detail_{version}")

# 결과를 텍스트 파일로 저장
count_diff = 0
if results:
    print("\n결과: ")
    for result in results:
        # 대상 파일명 추출, 저장할 텍스트 파일명 지정
        filename = result[6]
        output_filename = result_folder + f"\\{filename}.txt"
        if result[2] != [] or result[3] != [] or result[4] != [] or result[5] != []:
            # 실제로 유효한 결과가 존재할 때만 텍스트 파일로 저장
            with open(output_filename, "w", encoding="utf-8") as output_file:
                print(f"파일명: {result[0]}, 시트명: {result[1]} - 차이점 있음\n")
                output_file.write(f"파일명: {result[0]}, 시트명: {result[1]} - 차이점 있음\n\n")

                # 새 행이 추가됐을 때
                if result[4] != []:
                    print("[최신판에서 새로 추가된 행들]")
                    output_file.write("[최신판에서 새로 추가된 행들]\n")
                    for target in result[4]:
                        print(f"    해당 행 번호: {target[0]}")
                        output_file.write(f"    해당 행 번호: {target[0]}\n")
                        print(f"    해당 행 id: {target[1][0]}\n")
                        output_file.write(f"    해당 행 id: {target[1][0]}\n\n")
                        count_diff += 1

                # 이전 행이 제거됐을 때
                if result[5] != []:
                    print("[최신판에서 제거된 행들]")
                    output_file.write("[최신판에서 제거된 행들]\n")
                    for target in result[5]:
                        print(f"    해당 행 id: {target[1][0]}\n")
                        output_file.write(f"    해당 행 id: {target[1][0]}\n\n")
                        count_diff += 1

                # 차이가 발생한 행들에 대해
                if result[3] != []:
                    print("[최신판에서 변경된 행들]")
                    output_file.write("[최신판에서 변경된 행들]\n")

                    for r in range(0, len(result[2])):
                        # 하나의 row의 열수
                        length_row_curr = len(result[2][r][1])
                        length_row_prev = len(result[3][r][1])

                        length_row = min(length_row_curr, length_row_prev)

                        result_refined = []
                        # 정확히 어느 부분이 다른지 체크
                        for index in range(0, length_row):
                            if result[2][r][1][index] != result[3][r][1][index]:
                                result_refined.append((index, result[3][r][1][index], result[2][r][1][index]))

                        # 결과 출력
                        for rr in result_refined:
                            print(f"    해당 셀 행/열: [{result[2][r][0]}, {rr[0]}]")
                            print(f"    과거판 내용: {rr[1]}")
                            print(f"    최신판 내용: {rr[2]}\n")

                            output_file.write(f"    해당 셀 행/열: [{result[2][r][0]}, {rr[0]}]\n")
                            output_file.write(f"    과거판 내용: {rr[1]}\n")
                            output_file.write(f"    최신판 내용: {rr[2]}\n\n")
                            count_diff += 1

                # 결과 저장 알림
                print(f"작업 결과가 '{output_filename}'로 저장되었습니다.\n")

        else:
            print(f"파일명: {result[0]}, 시트명: {result[1]} - 차이점 없음")
else:
    print("검색 결과가 없습니다.")

# 최신판을 과거판으로 덮어씌울 것인가?
if count_diff != 0:
    print(f"\n비교 결과, 과거판과 비교하여 최신판에 {count_diff} 건의 변경점이 발견되었습니다.")
    print(f"검색 결과가 {result_folder} 경로에 저장되었습니다.")
else:
    print("\n비교 결과, 최신판은 과거판과 차이가 없습니다.")

yes = input("\n현재 최신판을 과거판으로 덮어씌우겠습니까? 작업을 완료한 뒤에 수행해 주십시오. (yes/n): ")
if yes == "yes" or yes == "YES":
    copy_tree(path_dir + "\\Mod_Korean", prev_folder + "\\Mod_Korean")

    print(f"\n비교 대상 파일을 최신화 하였습니다 (버전 {version}). 프로그램을 종료합니다.")
    exit()
else:
    exit()

# TODO
# 1) 일부 파일은 ASCII 알파벳으로 모자랄 정도로 열이 많다. 그냥 1행의 구분자로 바꾸는 게 나을듯 -> 보류