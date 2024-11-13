import os
from openpyxl import load_workbook

# 사용자로부터 검색할 텍스트 입력받기
search_text = input("검색할 텍스트를 입력하세요: ")

# 폴더 경로 설정
folder_path = "C:/Users/admin/Documents/GitHub/Elin_KR/_Lang_Korean/Lang/KR/Game"  # 엑셀 파일이 있는 폴더 경로

# 결과 저장 리스트
results = []

# 폴더 내 모든 파일 순회
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx") and not filename.startswith("~$"):  # ~$ 임시 파일 제외
        file_path = os.path.join(folder_path, filename)
        print(f"검색 중: {filename}")  # 현재 파일명 출력

        # 엑셀 파일 열기
        try:
            workbook = load_workbook(file_path, data_only=True)
        except PermissionError as e:
            print(f"권한 오류로 파일을 열 수 없습니다: {filename}")
            continue

        # 각 시트에서 검색
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 각 셀을 순회하면서 텍스트 검색
            found = False
            for row in sheet.iter_rows(values_only=True):
                if any(search_text in str(cell) if cell else False for cell in row):
                    results.append((filename, sheet_name))
                    found = True
                    break  # 텍스트가 발견되면 다음 시트로 이동

# 결과를 텍스트 파일로 저장
output_filename = f"검색결과_{search_text}.txt"  # 검색어를 포함한 파일명
with open(output_filename, "w", encoding="utf-8") as output_file:
    output_file.write(f"검색어: {search_text}\n\n")  # 검색어를 파일에 기록
    if results:
        for result in results:
            output_file.write(f"파일명: {result[0]}, 시트명: {result[1]}\n")
        print(f"검색 결과가 '{output_filename}'로 저장되었습니다.")
    else:
        output_file.write("검색 결과가 없습니다.\n")
        print("검색 결과가 없습니다.")
