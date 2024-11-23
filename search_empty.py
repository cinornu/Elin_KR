import os
import pathlib
from openpyxl import load_workbook
from datetime import datetime

# 폴더 경로 설정
# Elin_KR 폴더에 있는 것을 가정
path_dir = str(pathlib.Path(__file__).parent.resolve())
# Data 폴더 제외; 이 폴더의 파일들은 text_JP 등의 비교 구문이 없으므로 판별 불가
array_path = [path_dir + "\\Mod_Korean\\Lang\\KR\\Dialog", path_dir + "\\Mod_Korean\\Lang\\KR\\Dialog\\Drama", path_dir + "\\Mod_Korean\\Lang\\KR\\Game"]

# 검색할 열 지정자; 1행에 어떤 문자가 있느냐를 기준으로
col_identifier = ["text", "name", "detail", "textFlavor", "unit", "unknown", "roomName", "name2", "strPhase", "textPhase", "textPhase2", "textEnd", "textBenefit", "textType", "textAssign", "talkProgress", "talkComplete", "aka", "altName", "altname", "textExtra", "textInc", "textDec", "textAlt", "adjective", "levelBonus", "calm", "fov", "aggro", "dead", "kill"]

# 결과
results = []

# 제외할 파일명; 좌우에 text_JP 등의 비교문이 있는 파일에서만 제대로 작동함
search_skip = ["Backer.xlsx"]

# 하위 폴더들 내 모든 엑셀 파일 순회
for folder_path in array_path:
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx") and not filename.startswith("~$") and filename not in search_skip:  # ~$ 임시 파일 제외, 제외 리스트의 파일도 제외
            file_path = os.path.join(folder_path, filename)
            print(f"검색 중: {filename}")  # 현재 파일명 출력

        # 엑셀 파일 열기
            try:
                workbook = load_workbook(file_path, data_only=True)
            except PermissionError as e:
                print(f"권한 오류로 파일을 열 수 없습니다: {filename}")
                continue

            # 각 시트에서 검색
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 각 열 구분자에 맞춰서 반복
                for iden in col_identifier:
                # 1행에 identifier가 포함된 열 탐색
                    target_col = None
                    for col in sheet.iter_cols(1, sheet.max_column, 1, 1):
                        if col[0].value == iden:
                            target_col = col[0].column  # 열 번호 저장
                            break

                    if target_col is None:
                        print(f"시트 {sheet_name} 1행에 '{iden}' 셀이 없습니다.")
                        continue

                    # 해당 열에서 빈칸 탐색
                    for row in range(2, sheet.max_row + 1):  # 2행부터 탐색
                        # 빈칸 발견
                        if sheet.cell(row=row, column=target_col).value is None:
                            # 왼쪽/오른쪽 칸의 공백 여부로 "채워넣어야 하는" 빈칸 판별
                            # Dialog\Drama 디렉토리 - 그 왼쪽 칸이 빈칸이 아닐 경우
                            if "Drama" in folder_path and sheet.cell(row=row, column=target_col-1).value is not None:
                                # (파일 경로, 시트명, 열 구분자, 행 번호) 추가
                                results.append((file_path, sheet_name, iden, row))
                            # Game 디렉토리거나 Dialog 디렉토리 - 그 오른쪽 칸이 빈칸이 아닐 경우
                            elif sheet.cell(row=row, column=target_col+1).value is not None:
                                results.append((file_path, sheet_name, iden, row))

# 결과 출력
# 실행한 일자와 시간
now_time = datetime.now().strftime('%Y-%m-%d %H%M%S')

output_filename = f"공백검색결과_{now_time}.txt"
with open(output_filename, "w", encoding="utf-8") as output_file:
    output_file.write(f"공백 셀 검색 결과; 일자: {now_time}\n\n")
    if results:
        print("\n공백 셀 검색 결과: ")
        for result in results:
            print(f"파일명: {result[0]}, 시트명: {result[1]}, 위치: {result[2]}열, {result[3]}행")
            output_file.write(f"파일명: {result[0]}, 시트명: {result[1]}, 위치: {result[2]}열, {result[3]}행\n")
        print(f"검색 결과가 '{output_filename}'로 저장되었습니다.")
    else:
        output_file.write("검색 결과가 없습니다.\n")
        print("검색 결과가 없습니다.")